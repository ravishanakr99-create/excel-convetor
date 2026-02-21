"""Excel export: one row per PDF, one column per domain."""
import io
import logging
from typing import List, Dict, Any, Optional, Set
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows

logger = logging.getLogger(__name__)


def discover_fields(results: List[Dict[str, Any]]) -> List[str]:
    """Dynamically discover all fields from extraction results - pure alphabetical."""
    fields: Set[str] = set()
    for r in results:
        fields.update(r.get("fields", {}).keys())
    
    return sorted(fields)


def build_dataframe(
    results: List[Dict[str, Any]], 
    domain_columns: Optional[List[str]] = None
) -> pd.DataFrame:
    """Build DataFrame from extraction results: one row per PDF."""
    # Discover fields if not provided
    cols = domain_columns or discover_fields(results)
    
    rows = []
    for r in results:
        # Only file_name and extracted fields - no extra metadata
        row = {"file_name": r.get("file_name", "")}
        
        fields = r.get("fields", {})
        
        for col in cols:
            val = fields.get(col)
            row[col] = val if val is not None else ""
        
        rows.append(row)
    
    df = pd.DataFrame(rows)
    return df


def generate_excel(
    results: List[Dict[str, Any]],
    domain_columns: Optional[List[str]] = None,
    include_confidence: bool = False,
    include_summary: bool = False
) -> bytes:
    """
    Generate Excel file with extracted data.
    Each row = one PDF, each column = one field.
    """
    if not results:
        # Return empty Excel with headers
        df = pd.DataFrame(columns=["file_name", "status"])
    else:
        cols = domain_columns or discover_fields(results)
        df = build_dataframe(results, cols)
    
    if not include_confidence:
        drop_cols = [c for c in df.columns if c.endswith("_confidence")]
        df = df.drop(columns=drop_cols, errors="ignore")
    
    buffer = io.BytesIO()
    
    with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
        # Main data sheet
        df.to_excel(writer, sheet_name="Extracted Data", index=False)
        sheet = writer.sheets["Extracted Data"]
        
        # Style header
        for cell in sheet[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            cell.alignment = Alignment(horizontal="center", wrap_text=True)
        
        # Auto-adjust column widths
        for column in sheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            sheet.column_dimensions[column_letter].width = adjusted_width
        
        # Add summary sheet if requested
        if include_summary and results:
            summary_df = generate_summary_df(results)
            summary_df.to_excel(writer, sheet_name="Summary", index=False)
            summary_sheet = writer.sheets["Summary"]
            
            for cell in summary_sheet[1]:
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
    
    buffer.seek(0)
    return buffer.getvalue()


def generate_summary_df(results: List[Dict[str, Any]]) -> pd.DataFrame:
    """Generate summary statistics for the extraction."""
    total = len(results)
    successful = sum(1 for r in results if not r.get("error"))
    failed = total - successful
    scanned = sum(1 for r in results if r.get("is_scanned"))
    
    # Field coverage statistics
    all_fields = discover_fields(results)
    field_stats = []
    
    for field in all_fields:
        present = sum(1 for r in results if r.get("fields", {}).get(field))
        avg_confidence = 0
        confidences = [r.get("confidence_scores", {}).get(field, 0) for r in results]
        if confidences:
            avg_confidence = sum(confidences) / len(confidences)
        
        field_stats.append({
            "field_name": field,
            "extracted_count": present,
            "coverage_percent": round(100 * present / total, 1) if total > 0 else 0,
            "avg_confidence": round(avg_confidence, 2)
        })
    
    summary_data = [
        {"metric": "Total PDFs", "value": total},
        {"metric": "Successfully Processed", "value": successful},
        {"metric": "Failed", "value": failed},
        {"metric": "Scanned PDFs (OCR)", "value": scanned},
        {"metric": "Success Rate", "value": f"{100*successful/total:.1f}%" if total > 0 else "N/A"},
        {"metric": "", "value": ""},  # Empty row
        {"metric": "Field Coverage:", "value": ""},
    ]
    
    for stat in field_stats:
        summary_data.append({
            "metric": f"  {stat['field_name']}",
            "value": f"{stat['extracted_count']}/{total} ({stat['coverage_percent']}%) - avg conf: {stat['avg_confidence']}"
        })
    
    return pd.DataFrame(summary_data)
